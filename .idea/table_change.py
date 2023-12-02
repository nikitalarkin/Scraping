import openpyxl
import pandas as pd
# read_file = pd.read_csv (r'data/AlanyaHotels.xls', encoding="utf-8-sig")
# read_file.to_excel (r'data/AlanyaHotels.xlsx', index = None, header=True, encoding="utf-8-sig")
book = openpyxl.Workbook()
sheet = book.active
sheet["A1"] = 213
sheet.cell(row=1, column=1).value = 123
book.save('data\popa.xlsx')
sheet.cell(row=2, column=1).value = 'qweeqweq'
book.save('data\popa.xlsx')
book.close()