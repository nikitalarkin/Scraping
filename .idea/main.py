import requests, lxml
from bs4 import BeautifulSoup
import csv
import time
import openpyxl
from openpyxl.styles import Font

def get_data(url, city, page):
    headers = {
    "Accept": "text/javascript, application/javascript, application/ecmascript, application/x-ecmascript, */*; q=0.01",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "ru,en;q=0.9,en-GB;q=0.8,en-US;q=0.7",
    "Connection": "keep-alive",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.76"
    }
    r = requests.get(url=url, headers=headers)
    with open(f'indexes/index{city}page{page}.html', 'w',encoding="utf-8") as file:
        file.write(r.text)

urls = ['https://tophotels.ru/catalog/hotels/83?geo=15&df=2023-11-18&dt=2023-12-18&nf=7&nt=14&ad=2&ch=0&ch1=null&ch2=null&ch3=null&pf=0&pt=1000000&pc=0&text=&page=',
        'https://tophotels.ru/catalog/hotels/83?geo=18&df=2023-11-18&dt=2023-12-18&nf=7&nt=14&ad=2&ch=0&ch1=null&ch2=null&ch3=null&pf=0&pt=1000000&pc=0&text=&page=',
        'https://tophotels.ru/catalog/hotels/83?geo=14&df=2023-11-18&dt=2023-12-18&nf=7&nt=14&ad=2&ch=0&ch1=null&ch2=null&ch3=null&pf=0&pt=1000000&pc=0&text=&page=',
        'https://tophotels.ru/catalog/hotels/83?geo=17&df=2023-11-18&dt=2023-12-18&nf=7&nt=14&ad=2&ch=0&ch1=null&ch2=null&ch3=null&pf=0&pt=1000000&pc=0&text=&page=',
        'https://tophotels.ru/catalog/hotels/83?geo=16&df=2023-11-18&dt=2023-12-18&nf=7&nt=14&ad=2&ch=0&ch1=null&ch2=null&ch3=null&pf=0&pt=1000000&pc=0&text=&page='
        ]
cities = ['Antalya','Alanya', 'Kemer', 'Side', 'Belek']
pages_count = [0]*5
for i in range(0,5):
    with open(f'C:/my/programming/PyCharm/Projects/mama/indexes/index{cities[i]}.html', encoding="utf-8") as file:
        src = file.read()
    soup = BeautifulSoup(src, 'lxml')
    try:
        pages_count[i] = int(soup.find('a', class_='paginator__num fz13').text)
    except:
        pages_count[i] = 5
    book = openpyxl.Workbook()
    sheet = book.active
    aboba = ["Название","Населенный пункт","Линия","Звёзды","Рейтинг","Ссылка"]

    for j in range(6):
        sheet.cell(row=1, column=j+1).value = aboba[j]
        sheet.cell(row=1, column=j+1).font = Font(bold=True)
    sheet.auto_filter.ref = ("A1:E9999")
    sheet.column_dimensions['A'].width = 60
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['A'].width = 6
    sheet.column_dimensions['A'].width = 7
    sheet.column_dimensions['A'].width = 40


# for i in range(0,5): #(0,5)
#     for k in range(0, pages_count[i]): #(0, pages_count[i])  СОХРАНЕНИЕ СТРАНИЦ
#         get_data(f'{urls[i]}{k}', cities[i], k)
#         time.sleep(5)
    row = 2
    for k in range(0, pages_count[i]):
        with open(f'C:/my/programming/PyCharm/Projects/mama/indexes/index{cities[i]}page{k}.html', encoding="utf-8") as file:
            src = file.read()
        soup = BeautifulSoup(src, 'lxml')
        blocks = soup.find('div', class_='hotel-catalog-wrap').find_all(class_='hotel')
        for hotel in blocks:
            try:
                hotel_name = hotel.find('span', class_='hotel__name-cut').text
            except:
                break
            try:
                city = hotel.find('a', attrs={'data-item': 'filter-geo'}).find_next('a', attrs={'data-item': 'filter-geo'}).text.strip()[:-1]
            except:
                break
            try:
                line = hotel.find('a', attrs={'data-item': 'filter-pval'}).text.strip()
            except:
                line = '----'
            try:
                stars = float(hotel.find('span', attrs={'data-item': 'filter-cat'}).text[:-1])
            except:
                pass
            try:
                rating = float(hotel.find('span', class_='hotel__rate hotel__rate--green').find('b').text) #.replace('.', ',')
            except:
                pass
            try:
                rating = float(hotel.find('span', class_='hotel__rate hotel__rate--orange').find('b').text) #.replace('.', ',')
            except:
                pass
            try:
                rating = float(hotel.find('span', class_='hotel__rate hotel__rate--red').find('b').text) #.replace #('.', ',')
            except:
                pass
            href = hotel.find('a', target='_blank').get('href')

            sheet.cell(row=row, column=1).value = hotel_name
            sheet.cell(row=row, column=2).value = city
            sheet.cell(row=row, column=3).value = line
            sheet.cell(row=row, column=4).value = stars
            sheet.cell(row=row, column=5).value = rating
            sheet.cell(row=row, column=6).value = "https://tophotels.ru" + href
            row += 1
            book.save(f'data/{cities[i]}Hotels.xlsx')
    book.close()
    break